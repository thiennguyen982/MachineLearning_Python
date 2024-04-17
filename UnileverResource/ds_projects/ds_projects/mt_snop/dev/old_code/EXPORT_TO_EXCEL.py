# Databricks notebook source
# MAGIC %pip install xlsxwriter openpyxl

# COMMAND ----------

dbutils.widgets.text("BANNER_REMOVE_REPEATED", "")
dbutils.widgets.text("BANNER_PROMOTION", "")
dbutils.widgets.text("BANNER_NAME", "")
dbutils.widgets.text("USER_TRIGGER", "")

BANNER_REMOVE_REPEATED = dbutils.widgets.get("BANNER_REMOVE_REPEATED")
BANNER_PROMOTION = dbutils.widgets.get("BANNER_PROMOTION")
BANNER_NAME = dbutils.widgets.get("BANNER_NAME")
USER_TRIGGER = dbutils.widgets.get("USER_TRIGGER")

# BANNER_REMOVE_REPEATED = "2022_AEON REMOVE REPEATED FILE.parquet.gzip_Lai-Trung-Minh Duc.PDF"
# BANNER_PROMOTION = "2022_PROMOTION TRANSFORMATION_UPDATEDAEON.par_Lai-Trung-Minh Duc.PDF"
# BANNER_NAME = "AEON"
# USER_TRIGGER = "Lai-Trung-Minh.Duc@unilever.com"

for check_item in [BANNER_REMOVE_REPEATED, BANNER_PROMOTION, BANNER_NAME, USER_TRIGGER]:
  if check_item == "" or check_item == None:
    raise Exception("STOP PROGRAM - NO DATA TO RUN")
    
print (f"BANNER_REMOVE_REPEATED: {BANNER_REMOVE_REPEATED}")
print (f"BANNER_PROMOTION: {BANNER_PROMOTION}")
print (f"BANNER_NAME: {BANNER_NAME}")
print (f"USER_TRIGGER: {USER_TRIGGER}")


# COMMAND ----------

BANNER = BANNER_NAME
BANNER 

# COMMAND ----------

import pandas as pd 
import openpyxl
import xlsxwriter
import numpy as np
import os

from openpyxl import load_workbook
import shutil, os


# COMMAND ----------

PREDICT_FILE_PATH = f"/dbfs/mnt/adls/MT_POST_MODEL/RESULT/RESULT_{BANNER_NAME}.parquet"
TRAIN_FILE_PATH = f"/dbfs/mnt/adls/MT_POST_MODEL/RESULT/RESULT_{BANNER_NAME}.parquet"
PATH_PROMOTION_UPDATE = f"/dbfs/mnt/adls/MT_POST_MODEL/DATA_PROMOTION/UPDATE/"
#PATH_PROMOTION_HISTORY = f"/dbfs/mnt/adls/MT_POST_MODEL/DATA_PROMOTION/HISTORY/"

PROMOTION_FOLDER = f'/dbfs/mnt/adls/MT_POST_MODEL/PROMOTION_LIB/'
TEMPLATE_FILEPATH = f'/dbfs/mnt/adls/MT_POST_MODEL/MASTER/Template-Output-ML.xlsx'

OUTPUT_TEMPLATE = f'/dbfs/mnt/adls/MT_POST_MODEL/RESULT/'

START_DATA_FILE_LINE = 5
SHEETNAME = 'Result'

BANNER = [BANNER_NAME]

# COMMAND ----------

DF = pd.read_parquet(PREDICT_FILE_PATH)
DF.columns

# COMMAND ----------

def get_data_predict(file_path):
    df = pd.read_parquet(file_path)   
    df['REGION'] = 'total'
    df['FC']  = df['PREDICT']
    df['DEMAND SENSING (CS)'] = df['EST. DEMAND (CS)']
    if 'YEAR' not in df.columns:
      df['YEAR'] = df['YEARPOST']
    df['KEY'] = df.apply(lambda x: '{}|{}|{}|{}'.format(x['BANNER'],x['DP NAME'],x['POST NAME'],x['YEAR']),axis=1)
    df_total = df.groupby(['KEY']).agg({'DEMAND SENSING (CS)':'sum'}).reset_index()
    df_plant = pd.pivot_table(
        df,
        values=['DEMAND SENSING (CS)'],
        index=['KEY'],
        columns=['REGION'],
        aggfunc=sum,
        fill_value=0
    ).reset_index()

    df_plant.columns = df_plant.columns.map(' '.join).str.strip(' ')
    df = df_total.merge(df_plant, how='left', on='KEY')

    return df

# COMMAND ----------


def get_data_train(file_path): 
    df = pd.read_parquet(file_path)
    df = df.fillna(0)
    df['REGION'] = 'total'
    
    if 'YEAR' not in df.columns:
      df['YEAR'] = df['YEARPOST']    
    df['KEY'] = df.apply(lambda x: '{}|{}|{}|{}'.format(x['BANNER'],x['DP NAME'],x['POST NAME'],x['YEAR']),axis=1)
    df['BILLING (CS)'] = df['ACTUAL SALE (CS)']
    df['FC']  = df['PREDICT']
    df_plant = pd.pivot_table(
        df,
        values=['EST. DEMAND (CS)', 
           #'EST. DEMAND (CS)_CSE',
           'BILLING (CS)','FC'],
           #'ACTUAL SALE (CS)_CSE',FC],
        index=['KEY'],
        columns=['REGION'],
        aggfunc={
            'EST. DEMAND (CS)':np.sum, 
            'BILLING (CS)':np.sum,
           #'EST. DEMAND (CS)_CSE':np.sum,
           #'ACTUAL SALE (CS)_CSE':np.sum,
           'FC':np.sum
        },
        fill_value=0
    ).reset_index()

    df_plant.columns = df_plant.columns.map(' '.join).str.strip(' ')
    df ['CODE TYPE'] = df['TYPE']
    df ['% KA INVESTMENT (ON TOP)'] = 0
    df ['BASEKET PROMOTION'] = 0
    df['GROUP'] = 0
    df['MONTH'] = df['MONTH_TOTAL']
    df['WEEK'] = df['WEEK_TOTAL']
    df['BIG'] = 0
    df['SMALL'] = 0
    df['MEDIUM'] = 0
    df = df.groupby(['KEY','DP NAME', 'TYPE', 'POST NAME', 'CODE TYPE', 'SPECIAL_EVENT',
       'PROMOTION GROUPING', 'YEAR', 'GIFT TYPE', 'BASEKET PROMOTION',
       'ORDER END DATE', 'ORDER START DATE', 'POST_DISTANCE',
        'KEY_EVENT', 'CLASS', 'PROMOTION_IMPACT', 'ON/OFF POST', 'DAY',
       'PORTFOLIO (COTC)', 'PACKSIZE', 'PACKGROUP', 'BRAND VARIANT',
       'SUBBRAND', 'PRODUCTION SEGMENT', 'SEGMENT', 'FORMAT', 'CATEGORY',
       'BRAND','MTWORKINGSUM', 'DAY_BEFORE_HOLIDAY_CODE',
       'DAY_AFTER_HOLIDAY_CODE', 'COVID_IMPACT', 'MONTH_TOTAL', 'WEEK_TOTAL',
       'WEEKMONTH', 'DAY_TOTAL', '%TTS', '% BMI', 'GROUP',
       'PRICE CHANGE', '% KA INVESTMENT (ON TOP)', 'POST_LOST', 'MONTH',
       'WEEK', 'WEEKDAY', 'BIG', 'SMALL', 'MEDIUM', 'BANNER']).agg({
           'EST. DEMAND (CS)':'sum',
           'BILLING (CS)':'sum',  
           #'EST. DEMAND (CS)_CSE':'sum',
           #'ACTUAL SALE (CS)_CSE':'sum',
           'FC':'sum'
       }).reset_index()
    df = df.merge(df_plant, how='left', on='KEY')
    return df

# COMMAND ----------

def get_data_promo(fileitem, folder):
    print(fileitem)
    df = pd.read_excel(folder + fileitem, sheet_name = 'Promotion library',skiprows=1,engine='openpyxl')
    df = df.drop(df.index[[0,1,2,3]])
    df = df[df['Banner'].isin(BANNER)]
    return df

def promotion_tranform():
    print('Read promotion')
    df = pd.DataFrame()
    for fileitem in os.listdir(PROMOTION_FOLDER):
        df_temp = get_data_promo(fileitem, PROMOTION_FOLDER)
        df = pd.concat([df,df_temp])
    for fileitem in os.listdir(PATH_PROMOTION_UPDATE):
        df_temp = get_data_promo(fileitem, PATH_PROMOTION_UPDATE)
        df = pd.concat([df,df_temp])

    for i in ['Banner','DP Name','Post Name']:
        df[i]=df[i].str.lower()

    df['KEY'] = df.apply(lambda x: '{}|{}|{}|{}'.format(x['Banner'],x['DP Name'],x['Post Name'],x['Year']),axis=1)

    for i in ['Original FC (cs)',
        'Final allocation (cs)',
        'Original allocation (cs) V102',
        'Original allocation (cs) V103',
        'Original allocation (cs) V101',
        'Final allocation (cs) V102',
        'Final allocation (cs) V103',
        'Final allocation (cs) V101']:
        df[i] = pd.to_numeric(df[i],errors='coerce')

    df = df.groupby(['KEY']).agg({
        'Original FC (cs)':'sum',
        'Final allocation (cs)':'sum',
        'Original allocation (cs) V102':'sum',
        'Original allocation (cs) V103':'sum',
        'Original allocation (cs) V101':'sum',
        'Final allocation (cs) V102':'sum',
        'Final allocation (cs) V103':'sum',
        'Final allocation (cs) V101':'sum',
        }).reset_index()

    return df

# COMMAND ----------

def get_data_todict():
    # df_predict = get_data_predict()
    # df = get_data_train()
    df = get_data_train(TRAIN_FILE_PATH) 
    df['BANNER'] = df['BANNER'].str.upper()
    
    df_promo = promotion_tranform()
    # df = df_train.merge(df_predict,how='left',on='KEY')
    df['DEMAND SENSING (CS)'] = df['FC'].fillna(df['FC'])
    df = df.merge(df_promo,how='left',on='KEY')
    df = df.sort_values(by=['ORDER START DATE'],ascending = True)
    df = df.sort_values(by=['DP NAME'])
    df = df.reset_index(drop=True)
    df['No'] = 1
    df['No'] = df.index

    df_dict = df.to_dict('list')
    return df_dict

# COMMAND ----------

def get_column_location(header, header_in_worksheet):
    for header_location in header_in_worksheet:
        header_name = header_location.value
        column_location = header_location.column
        if (header_name == header):
            return column_location

def add_data_to_template(df_dict):
    print ('Load Template Promotion')
    workbook = load_workbook(TEMPLATE_FILEPATH)
    worksheet = workbook[SHEETNAME]
    header_in_worksheet = worksheet[3]

    length_of_dataset = len(df_dict['BANNER'])

    print ('Load Data From DataFrame to Template')
    for header_location in header_in_worksheet:
        header_name = header_location.value
        column_location = header_location.column

        if (header_name in df_dict.keys()):    
            for row in range(0, length_of_dataset):
                cell_value = df_dict[header_name][row]
                try:
                    worksheet.cell(row=START_DATA_FILE_LINE + row, column=column_location, value=cell_value)
                except:
                    pass
              
    return workbook

# COMMAND ----------

df_dict = get_data_todict()
workbook = add_data_to_template(df_dict)


# COMMAND ----------

name = "/tmp/" + "MT_POST_" + BANNER_NAME+".xlsx"
name_out = "MT_POST_" + BANNER_NAME+'.xlsx'
writer = pd.ExcelWriter( name ,engine='xlsxwriter')
workbook.save(writer)
shutil.copy(name, OUTPUT_TEMPLATE)


# COMMAND ----------

OUTPUT_TEMPLATE

# COMMAND ----------

import requests
import base64
import time

url = "https://prod-215.westeurope.logic.azure.com:443/workflows/72d08b2ad319428f9927758f1d87f892/triggers/manual/paths/invoke?api-version=2016-06-01&sp=%2Ftriggers%2Fmanual%2Frun&sv=1.0&sig=erridJccQ8dln3JknLJxgbgnR55TCPDrNwqdWrd10CU"

with open(
    f"/dbfs/mnt/adls/MT_POST_MODEL/RESULT/MT_POST_{BANNER_NAME}.xlsx",
    "rb",
) as my_file:
    encoded_string = base64.b64encode(my_file.read()).decode("utf-8")

my_obj = {
    "KEY": f"RESULT_{BANNER_NAME}_{int(time.time())}.xlsx",
    "USER": USER_TRIGGER,
    "BANNER": BANNER_NAME,
    "FILEINDEX_CONTENT": encoded_string,
}

requests.post(url, json=my_obj)

# COMMAND ----------

